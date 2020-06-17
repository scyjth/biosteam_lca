# -*- coding: utf-8 -*-
"""
Created on Thu Apr 18 14:23:21 2019

@author: cyshi
"""
from __future__ import unicode_literals
from ._unicode import UnicodeReader
import csv
import xlrd
from voluptuous import Schema, Required, Invalid, Any, All, Length, Optional
from numbers import Number
import os 
import warnings
from numbers import Number
import numpy as np
from .import methods, strategies, config, importers, Database
import operator
import functools 
import sys
import time
import openpyxl
from tabulate import tabulate
from bw2data.ia_data_store import ImpactAssessmentDataStore

PY3 = sys.version_info > (3,)
if PY3:
    dt = ";"
else:
    dt = b";"

dirpath = os.path.dirname(__file__)

def valid_tuple(obj):
    """To validate if a impact assesment name is a valid key tuple. Ia_method should follow the format such as
    ('TRACI', 'environmental impact', 'acidification')
    """
    try:
        assert isinstance(obj, tuple)
        assert isinstance(obj[0], str)
        assert isinstance(obj[1], str)
    except:
        raise Invalid("{} is not a valid key tuple".format(obj))
    return obj


uncertainty_dict = {
    Required("amount"): Number,
    Optional("uncertainty type"): int,
    Optional("loc"): Number,
    Optional("scale"): Number,
    Optional("shape"): Number,
    Optional("minimum"): Number,
    Optional("maximum"): Number}

exchange = {Required("input"): valid_tuple, Required("type"): str,}
exchange.update(uncertainty_dict)

maybe_uncertainty = Any(Number, uncertainty_dict)
validator_ia = Schema([Any([valid_tuple, maybe_uncertainty],[valid_tuple, maybe_uncertainty, object])])

class ImpactAssessmentMethods():
    
    @staticmethod
    def search (search_string = None, exclude = None, length = None):
            """
             A method to lookup life cycle impact assessment method (LCIA) by method name or impact category, such as:'IPCC 2007', or 'ReCiPe Endpoint (H,A)'
             By tradition, LCIA method is understood as a set of LCIA impact catergories, it need to be selected before caulcating LCA results for product, system, or projectis.
             Returns all methods for which the search string is part of. If no mthod is specified, will return US EPA TRACI.
             Certain methods will be exclude if exclude string is specified.  
             If length is specified, results are reduced to those that have a length of x, i.e. x parts in the method tuple.
            """
            asmethod = []
            if search_string == None:
                asmethod = [
                    ('TRACI', 'environmental impact', 'acidification'),
                     ('TRACI', 'environmental impact', 'ecotoxicity'),
                     ('TRACI', 'environmental impact', 'eutrophication'),
                     ('TRACI', 'environmental impact', 'global warming'),
                     ('TRACI', 'environmental impact', 'ozone depletion'),
                     ('TRACI', 'environmental impact', 'photochemical oxidation'),
                     ('TRACI', 'human health', 'carcinogenics'),
                     ('TRACI', 'human health', 'non-carcinogenics'),
                     ('TRACI', 'human health', 'respiratory effects, average')]        
            else:
                #asmethod = [method for method in methods if search_string in method[0] ]
                for method in methods:
                    for name in method:
                        if search_string.lower() in name.lower() :
                            asmethod.append(method)
            if exclude:
                asmethod = [op for op in asmethod if exclude not in [sth for sth in op]]
            if length:
                asmethod = [op for op in asmethod if len(op) == length]
            if not asmethod:
                raise ValueError("No methods matches your selection criteria. Please select a valide life cycle impact assessment method")   
            return asmethod

    def categories(asmethods=None):
        """Analyze impact categories.
        Args:
            * *asmethods*: List of candidate impact assessment methods
        """
        if asmethods == None:
            asmethods = methods
        ct_dict={}
        am=[method[1] for method in asmethods]
        for i in am:
            am.count(i)
            ct_dict[i]=am.count(i)
        sorted_ct = sorted(ct_dict.items(), key=operator.itemgetter(1))
        return sorted_ct
    
    def tools(asmethods): 
        """
        Summerize total approaches. 
        Eco-Indicator 99, EPS Method, LIME, and Impact 2002+ include potential impacts from future extractions in the impact assessment, \
        whereas Recipe assumes assumes impacts have been included in the inventory analysis.
        ReCipe method includes eighteen midpoint indicators and three endpoint indicators
        
        **Args**
            * *asmethods* (list): List of candidate impact assessment methods
        """
        tm=[method[0] for method in asmethods]
        tools_lst = list(dict.fromkeys(tm))
        return tools_lst, ("Number of methods: {}".format(len(tools_lst)))
    
    def negtive_cf(method_name=None):
        """find the negative values of Chracterization Fators of ia method, which implied that they are good for the environment
        **kwargs    
            **method_name** : [str] name of the method tool, for exmaple-- ('ReCiPe Midpoint (I) V1.13')
        **return
            A dictionary of negtive CFs of selected method. For exmaple-- 
            ('ReCiPe Midpoint (I) V1.13','photochemical oxidant formation','POFP'):\ 
                [('Benzaldehyde' (kilogram, None, ('air',)),-0.155),
                ('Benzaldehyde' (kilogram, None, ('air', 'urban air close to ground')), -0.155)]
        """
        negtive_cf_dict={}
        if method_name is not None:
            method_lst= list(filter(lambda x: x[0] == method_name, methods))
        else:
            method_lst=methods
        for method in method_lst:
            CFs ={}
            for key, cfs in Method(method).load():
                if cfs < 0:
                    CFs[((Database('biosphere3').get(key[1])))] = cfs
            if CFs:
                negtive_cf_dict[method]= CFs
#        workbook = openpyxl.Workbook()
#        sheet = workbook.active
#        row = 1
#        for key,values in negtive_cf_dict.items():
#            sheet.cell(row=row, column=1, value=str(key))
#            column = 2
#            for element in values:
#                sheet.cell(row=row, column=column, value=str(element))
#                column += 1
#                row += 1
#        workbook.save(filename="ncf2.xlsx")
        return negtive_cf_dict
        
    def update_lcia_methods(overwrite=True):
        start = time.time()
        ei = SetLCIA()
        ei.apply_strategies(verbose=False)
        ei.write_methods(overwrite=overwrite)
        end = time.time()
        print("writing time:".format(end - start))
    
    def overview (self, list_all=False):
            return list(methods)

    def __str__(self):
        return "%s" % (self.__class__.__name__) 

LCIAImporter = importers.base_lcia.LCIAImporter 

class SetLCIA(LCIAImporter):
    """Create and/or update lcia methods. A dictionary for method metadata. File data is saved in ``methods.json``.
    filename = "methods.json". See data_serialization
    lcia_methods in ecoinvent can be found at "http://www.ecoinvent.org/support/documents-and-files/information-on-ecoinvent-3/information-on-ecoinvent-3.html"
    """

    def __init__(self):
        self.strategies = [strategies.normalize_units, strategies.set_biosphere_type, 
                           strategies.drop_unspecified_subcategories,
                           functools.partial(strategies.link_iterable_by_fields,
                                             other = Database(config.biosphere),
                                             fields = ('name', 'categories')),]
        self.applied_strategies = []
        self.csv_data, self.cf_data, self.units, self.file = self.lcia_methods__metadata()
        self.separate_methods()
        
    def lcia_methods__metadata(self):
        """convert impact assessment data. Raw data is saved in the "lcia_implemnetation_2019" file"""
        with UnicodeReader(os.path.join(dirpath, "categoryUUIDs.csv"), 
                           encoding='latin-1', 
                           delimiter=dt) as csv_file:
            next(csv_file) 
            csv_data = [{'name': (line[0], line[2], line[4]),
                'description': line[7]
            } for line in csv_file]
    
        filename = "LCIA_implementation_2019.xlsx" # this was donwloaded and updated on Oct 2019 from ecoinvent website. 
        wb = xlrd.open_workbook(os.path.join(dirpath, filename))
        #characterizaton factors
        sheet= wb.sheet_by_name("CFs")
        cf_data = [{
            'method': (sheet.cell(row, 0).value,
                       sheet.cell(row, 1).value,
                       sheet.cell(row, 2).value),
            'name': sheet.cell(row, 3).value,
            'categories': (sheet.cell(row, 4).value, sheet.cell(row, 5).value),
            'amount': sheet.cell(row, 7).value
            }
            for row in range(1, sheet.nrows)
            if sheet.cell(row, 0).value not in 
            {'selected LCI results, additional', 'selected LCI results'} and isinstance(sheet.cell(row, 7).value, Number)]
        #units
        sheet= wb.sheet_by_name("units")
        units = {(sheet.cell(row, 0).value, sheet.cell(row, 1).value, 
                  sheet.cell(row, 2).value): sheet.cell(row, 4).value for row in range(1, sheet.nrows)}
        return csv_data, cf_data, units, filename

    def separate_methods(self):
        """Separate the list of CFs into distinct methods"""
        methods = {obj['method'] for obj in self.cf_data}
        metadata = {obj['name']: obj for obj in self.csv_data}
        self.data = {}
        missing = set()
        for line in self.cf_data:
            if line['method'] not in self.units:
                missing.add(line['method'])

        if missing:
            _ = lambda x: sorted([str(y) for y in x])
            warnings.warn("Missing units for following:" +
                             " | ".join(_(missing)))

        for line in self.cf_data:
            assert isinstance(line['amount'], Number)

            if line['method'] not in self.data:
                self.data[line['method']] = {
                    'filename': self.file,
                    'unit': self.units.get(line['method'], ''),
                    'name': line['method'],
                    'description': '',
                    'exchanges': []}
            self.data[line['method']]['exchanges'].append({
                'name': line['name'],
                'categories': line['categories'],
                'amount': line['amount']})

        self.data = list(self.data.values())

        for obj in self.data:
            obj.update(metadata.get(obj['name'], {}))  
           
class Method (ImpactAssessmentDataStore):
    
    _metadata = methods

    def __init__(self, name):
        self.name = name

    def _write(self, data, process=True):
        """Serialize intermediate data to disk.

        Sets the metadata key ``num_cfs`` automatically."""
        self.metadata[u"num_cfs"] = len(data)
        self._metadata.flush()
        super(Method, self).write(data)
        
    def characterization_factor(self):
        """A method to find characterization factors of a lcia method"""
        CFs=[]
        for key, cf in super(Method, self).load():
            try:
                activity= Database(key[0]).get(key[1])
            except TypeError as err:
                print (err)
            CFs.append((activity, cf))
        return CFs
    
    def get_abbreviation(self):
        """Retrieve the abbreviation of the method identifier."""
        self.register()
        return self.metadata["abbreviation"]
            
    def description(self):
        return self.metadata.get('description')
    
    def unit(self):
        return self.metadata.get('unit')
        
    def __str__(self):
        return "Impact Assessment %s: %s in BioSteam.LCA" % (self.__class__.__name__, self.name)            
      
Methods = ImpactAssessmentMethods