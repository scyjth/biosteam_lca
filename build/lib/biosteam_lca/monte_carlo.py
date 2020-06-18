# -*- coding: utf-8 -*-
"""
Created on Wed May 29 16:44:54 2019

@author: cyshi
"""
import uuid
import openpyxl
import pyprind
import stats_arrays as sa
import brightway2 as bw2
from multiprocessing import cpu_count
#from databases import SetUp
import numpy as np
import matplotlib.pyplot as plt
#from .method_constructor import Method
##THIS IS A DRAFT VERSION


def uncertainify(data, distribution=None, bounds_factor=0.1, sd_factor=0.1):
    """
    Add rough uncertainty to exchanges. This function only apply to exchanges with no uncertainty type or uncertainty type ``UndefinedUncertainty``. This does not change production exchanges!
    
    Can only apply normal or uniform uncertainty distributions; default is uniform distribution, if specified, must be a ``stats_array`` uncertainty object.
    
    ``data`` is a LCI data dictionary.
    
    If using the normal distribution:
    
    * ``sd_factor`` will be multiplied by the mean to calculate the standard deviation.
    * If no bounds are desired, set ``bounds_factor`` to ``None``.
    * Otherwise, the bounds will be ``[(1 - bounds_factor) * mean, (1 + bounds_factor) * mean]``.
    
    If using the uniform distribution, then the bounds are ``[(1 - bounds_factor) * mean, (1 + bounds_factor) * mean]``.
    Returns the modified data.
    """
    assert distribution in {None, sa.UniformUncertainty, sa.NormalUncertainty}, u"``uncertainify`` only supports normal and uniform distributions"
    assert bounds_factor is None or bounds_factor * 1. > 0, "bounds_factor must be a positive number"
    assert sd_factor * 1. > 0, "sd_factor must be a positive number"
    for key, value in data.items():
        for exchange in value.get(u'exchanges', []):
            if (exchange.get(u'type') == u'production') or \
                    (exchange.get(u'uncertainty type', sa.UndefinedUncertainty.id) \
                    != sa.UndefinedUncertainty.id):
                continue
            if exchange[u"amount"] == 0:
                continue
            if bounds_factor is not None:
                exchange.update({ u"minimum": (1 - bounds_factor) * exchange['amount'],u"maximum": (1 + bounds_factor) * exchange['amount'],})
                if exchange[u"amount"] < 0:
                    exchange[u"minimum"], exchange[u"maximum"] = exchange[u"maximum"], exchange[u"minimum"]
            if distribution == sa.NormalUncertainty:
                exchange.update({
                    u"uncertainty type": sa.NormalUncertainty.id,
                    u"loc": exchange[u'amount'],
                    u"scale": abs(sd_factor * exchange[u'amount']),})
            else:
                assert bounds_factor is not None, "must specify bounds_factor for uniform distribution"
                exchange.update({u"uncertainty type": sa.UniformUncertainty.id,})
    return data

class MultiMonteCarlo(object):
    """MonteCarlo analysis using multiple impact assessment methods.
    Args:
        * ``args`` is a list of LCIA methods
        * ``method`` is a lci flow
        * ``iterations`` is the number of Monte Carlo iterations desired
    
    Call ``.uncertainty()`` to generate results.    
    """
    
    def __init__(self, flow, methods_array):
        self.flow = flow
        self.methods_array = methods_array
        #self.units = list(map(lambda x: Method(x).unit(), self.methods_array))
    
    def uncertainty(self, iteration=100,plot=True):
        uncertainty_array={}
        for m in self.methods_array:
            flow_mc = bw2.MonteCarloLCA({self.flow:1}, method=m)
            scores = [next(flow_mc) for _ in pyprind.prog_bar(range(iteration))]
            if plot==True:
                print (m)
                # plt.plot()
                plt.figure(self.methods_array.index(m))
                plt.hist(np.array(scores), histtype='step', bins=100, density=True)
                #plt.xlabel("{}".format(Method(m).metadata['unit']))
                plt.ylabel("Frequency")
                plt.legend()
                plt.title("Monte Carlo results\n{}\nUnit processes".format(m))    
            uncertainty_array[m] = np.array(scores)
            plt.show()
        return uncertainty_array
    
    def to_excel(self, dict_name, file_name = "mc_lci_workbook.xlsx"):
        self.dict_name = dict_name
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # openpyxl does things based on 1 instead of 0
        row = 1
        for key,values in self.dict_name.items():
            # Put the key in the first column for each key in the dictionary
            sheet.cell(row=row, column=1, value=key)
            column = 2
            for element in values:
                # Put the element in each adjacent column for each element in the tuple
                sheet.cell(row=row, column=column, value=element)
                column += 1
            row += 1      
        workbook.save(filename = file_name)
    def __str__(self):
        return "%s" % (self.__class__.__name__)           


# LCA_calc_mc = {} #store lca data
# def monte_carlo (flow, method, amount, iterations=500, cpus = None, uuid_=None):
#     if cpus == None:
#         cpus = cpu_count()
#     mc_data = SerializedLCAReport({flow: amount}, method, iterations, cpus).get_monte_carlo()
#     if uuid_:
#         mc_data['iterations'] = iterations
#         LCA_calc_mc.update({uuid_: mc_data})
#     return mc_data


#test 
#uuid_ = str(uuid.uuid4().urn[9:])
#a=SetUp()
#flow = a.set_input('electricity')[0]
#im_methods=a.set_method('CML')[0]
#m =monte_carlo(flow, im_methods, amount = 1)   